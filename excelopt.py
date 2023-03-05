from openpyxl import load_workbook


def get_execl_info(file_name, start_row=2, end_row=9999):
    """
    :param file_name: 文件名
    :param start_row:  开始行
    :param end_row: 结束行
    :return:  返回一个从start_row 到 end_row的列表,列表中包含每一行的信息
    """
    infos = []
    wb = load_workbook(file_name)
    ws = wb.active
    end_row = min(ws.max_row+1, end_row+1)
    max_col = ws.max_column
    for row in range(start_row, end_row):
        row_info = []
        for col in range(max_col):
            row_info.append(ws[row][col].value)
        infos.append(row_info)
    return infos


if __name__ == '__main__':
    for line in get_execl_info('账户添加表.xlsx',2,10):
        print(line)

