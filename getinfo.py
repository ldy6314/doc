from openpyxl import load_workbook, Workbook

wb = load_workbook('bh.xlsx')
sn = wb.sheetnames[0]
ws = wb[sn]
mr = ws.max_row
mc = ws.max_column
out = Workbook()
ob = out.active

place = {
    'C1': '46中',
    'C2': '附小南校区',
    'C3': '望湖小学西区',
    'C4': '48滨湖校区',
    'X1': '46中',
    'X2': '附小南校区',
    'X3': '望湖小学西区',
    'X4': '48滨湖',
    'X5': '青小香港街校区',
    'X6': '卫岗小学',

}
infos = []
for row in range(3, mr + 1):
    num = ws[row][7].value
    pl = num[2:4]
    kc = num[4]
    zw = num[5:]
    infos.append((place[pl], kc, zw))

for i in infos:
    ob.append(i)

out.save('信息补充.xlsx')